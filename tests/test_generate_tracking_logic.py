import pandas as pd
from openpyxl import load_workbook

from tabs.generate_tracking.comparison_logic import aggregate_unique, build_template_sheet_df, classify_new_old
from tabs.generate_tracking.excel_writer.three_uk_qualys import (
    THREE_UK_QUALYS_TOTAL_COLUMNS,
    build_3uk_qualys_template_sheet_df,
)
from tabs.generate_tracking.excel_writer.workbook import write_output
from tabs.generate_tracking.parser import TEMPLATE_COLUMNS


def test_classify_new_old_uses_selected_raw_rows_and_master_key_presence():
    raw = pd.DataFrame(
        [
            {"Title": "Apache Vuln", "IP": "10.0.0.1", "Port": "443", "CVE ID": "CVE-1111", "Raw Only": "raw-a"},
            {"Title": "Nginx Vuln", "IP": "10.0.0.3", "Port": "80", "CVE ID": "CVE-3333", "Raw Only": "raw-n"},
            {"Title": "OpenSSL Vuln", "IP": "10.0.0.2", "Port": "443.0", "CVE ID": "CVE-2222", "Raw Only": "raw-o"},
            {"Title": "", "IP": "", "Port": "", "CVE ID": "", "Raw Only": "raw-blank"},
        ]
    )
    master = pd.DataFrame(
        [
            {"Name": " apache vuln ", "Host / Image": "10.0.0.1", "Port": 443, "CVE": "cve-1111"},
            {"Name": "OpenSSL Vuln", "Host / Image": "10.0.0.2", "Port": "443", "CVE": "CVE-2222"},
            {"Name": "", "Host / Image": "", "Port": "", "CVE": ""},
        ]
    )

    new_df, old_df = classify_new_old(raw, master)

    assert new_df["Title"].tolist() == ["Nginx Vuln", ""]
    assert old_df["Title"].tolist() == ["Apache Vuln", "OpenSSL Vuln"]
    assert old_df["Raw Only"].tolist() == ["raw-a", "raw-o"]


def test_classified_raw_rows_are_rendered_in_template_columns():
    raw = pd.DataFrame(
        [
            {"Plugin Name": "Apache Vuln", "Host": "10.0.0.1", "Service Port": "443", "CVE IDs": "CVE-1111", "Risk Factor": "High"},
            {"Plugin Name": "Nginx Vuln", "Host": "10.0.0.3", "Service Port": "80", "CVE IDs": "CVE-3333", "Risk Factor": "Medium"},
        ]
    )
    master = pd.DataFrame(
        [{"Name": "Apache Vuln", "Host / Image": "10.0.0.1", "Port": "443", "CVE": "CVE-1111"}]
    )

    new_df, old_df = classify_new_old(raw, master)
    template_new_df = build_template_sheet_df(new_df)
    template_old_df = build_template_sheet_df(old_df)

    assert template_new_df.columns.tolist() == TEMPLATE_COLUMNS
    assert template_old_df.columns.tolist() == TEMPLATE_COLUMNS
    assert template_new_df.loc[0, "Name"] == "Nginx Vuln"
    assert template_new_df.loc[0, "Risk"] == "Medium"
    assert template_old_df.loc[0, "Name"] == "Apache Vuln"
    assert template_old_df.loc[0, "Risk"] == "High"


def test_aggregate_unique_merges_all_raw_rows_preserving_order():
    df = pd.DataFrame(
        [
            {**{col: "" for col in TEMPLATE_COLUMNS}, "Name": "A", "CVE": "CVE-1", "Host / Image": "h", "Port": "80", "Risk": "Medium"},
            {**{col: "" for col in TEMPLATE_COLUMNS}, "Name": "A", "CVE": "CVE-1", "Host / Image": "h", "Port": "443", "Risk": "Critical"},
            {**{col: "" for col in TEMPLATE_COLUMNS}, "Name": "A", "CVE": "CVE-1", "Host / Image": "h", "Port": "80", "Risk": "Low"},
        ]
    )

    unique_df = aggregate_unique(df)

    assert len(unique_df) == 1
    assert unique_df.loc[0, "Port"] == "80, 443"
    assert unique_df.loc[0, "Risk"] == "Critical"


def test_3uk_qualys_output_keeps_total_in_qualys_format(tmp_path):
    total_df = pd.DataFrame([{col: "" for col in THREE_UK_QUALYS_TOTAL_COLUMNS}])
    total_df.loc[0, "IP"] = "10.0.0.1"
    total_df.loc[0, "QID"] = "123"
    total_df.loc[0, "Title"] = "Apache Vuln"
    total_df.loc[0, "Port"] = "443"
    total_df.loc[0, "CVE ID"] = "CVE-1111"
    total_df.loc[0, "Criticality"] = "HIGH"
    template_df = build_3uk_qualys_template_sheet_df(total_df)
    path = tmp_path / "tracking.xlsx"

    write_output(path, template_df.iloc[0:0], template_df, template_df, "3UK", "Qualys", total_df=total_df)

    wb = load_workbook(path, read_only=True)
    assert wb.sheetnames == [
        "Dashboard",
        "Total Vulnerabilities",
        "New Vulnerabilities",
        "Old Vulnerabilities",
        "Unique Vulnerabilities",
        "Disposition",
    ]
    ws_total = wb["Total Vulnerabilities"]
    assert [ws_total.cell(row=2, column=i).value for i in range(1, 4)] == ["IP", "Network", "DNS"]
    assert ws_total.cell(row=3, column=1).value == "10.0.0.1"
    wb.close()
