"""Shared Excel styling and sizing helpers."""

from __future__ import annotations

from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)

from openpyxl.utils import get_column_letter

from tabs.make_new_report.parser import TEMPLATE_COLUMNS


# ---------------------------------------------------------
# COLORS
# ---------------------------------------------------------

BLUE = PatternFill(
    start_color="001D3F72",
    end_color="001D3F72",
    fill_type="solid",
)

GREEN = PatternFill(
    start_color="008BC34A",
    end_color="008BC34A",
    fill_type="solid",
)


# ---------------------------------------------------------
# FONTS
# ---------------------------------------------------------

WHITE = Font(
    color="FFFFFF",
    bold=True,
)

BOLD = Font(
    bold=True,
)


# ---------------------------------------------------------
# ALIGNMENTS
# ---------------------------------------------------------

CENTER = Alignment(
    horizontal="center",
    vertical="center",
)

DATA_ALIGNMENT = Alignment(
    wrap_text=True,
    vertical="top",
    horizontal="right",
)


# ---------------------------------------------------------
# THIN BLACK BORDER
# ---------------------------------------------------------

BLACK_THIN_BORDER = Border(
    left=Side(
        style="thin",
        color="000000",
    ),
    right=Side(
        style="thin",
        color="000000",
    ),
    top=Side(
        style="thin",
        color="000000",
    ),
    bottom=Side(
        style="thin",
        color="000000",
    ),
)


# ---------------------------------------------------------
# META ROW
# ---------------------------------------------------------

def style_meta_row(
    ws,
    total_cols: int = len(TEMPLATE_COLUMNS),
):

    for col in range(1, total_cols + 1):

        cell = ws.cell(
            row=1,
            column=col,
        )

        cell.fill = BLUE
        cell.font = WHITE
        cell.alignment = CENTER
        cell.border = BLACK_THIN_BORDER


# ---------------------------------------------------------
# HEADERS
# ---------------------------------------------------------

def write_headers(
    ws,
    columns: list[str] = TEMPLATE_COLUMNS,
):

    ws["A1"] = "Project Name:"
    ws["D1"] = "Scanner:"

    style_meta_row(
        ws,
        total_cols=len(columns),
    )

    for i, col_name in enumerate(columns, start=1):

        display_name = (
            "Host / Image"
            if col_name == "Host"
            else col_name
        )

        cell = ws.cell(
            row=2,
            column=i,
            value=display_name,
        )

        if i <= 14:

            cell.fill = GREEN
            cell.font = BOLD

        else:

            cell.fill = BLUE
            cell.font = WHITE

        cell.alignment = CENTER
        cell.border = BLACK_THIN_BORDER


# ---------------------------------------------------------
# SUMMARY HEADERS
# ---------------------------------------------------------

def write_summary_headers(
    ws,
    project: str,

    scanner: str,
):

    ws["A1"] = "Project Name:"
    ws["D1"] = "Scanner:"

    ws["B1"] = project
    ws["E1"] = scanner

    style_meta_row(ws)


# ---------------------------------------------------------
# APPLY FULL FORMATTING
# ---------------------------------------------------------

def apply_table_formatting(ws, include_borders: bool = True):

    for row in ws.iter_rows():

        for cell in row:

            if include_borders:
                cell.border = BLACK_THIN_BORDER

            # Header rows
            if cell.row in [1, 2]:

                cell.alignment = CENTER

            else:

                cell.alignment = DATA_ALIGNMENT

    # -----------------------------------------------------
    # FIXED ROW HEIGHT
    # -----------------------------------------------------

    for row_idx in range(1, ws.max_row + 1):

        ws.row_dimensions[row_idx].height = 15


# ---------------------------------------------------------
# AUTO WIDTH
# ---------------------------------------------------------

def auto_width(ws):

    for col in ws.columns:

        idx = col[0].column

        max_len = max(
            (
                len(str(c.value))
                for c in col
                if c.value is not None
            ),
            default=10,
        )

        ws.column_dimensions[
            get_column_letter(idx)
        ].width = min(
            max(12, max_len + 4),
            45,
        )
