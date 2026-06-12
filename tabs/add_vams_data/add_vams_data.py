"""Tab 3: merge VAMS data into an existing generated workbook."""

import customtkinter as ctk
import pandas as pd


from openpyxl import (
    load_workbook,
)

from tabs.add_vams_data.fast_vams_enrichment import (
    FastVamsEnrichmentEngine,
    build_fast_keys,
)

from tabs.add_vams_data.excel_writer.reader import read_sheet_as_df
from tabs.add_vams_data.excel_writer.sheets import write_summary_sheet

from tabs.add_vams_data.parser import parse_scan_file

from tabs.add_vams_data.logic import (
    TOTAL_SHEET_CANDIDATES,
    UNIQUE_SHEET_CANDIDATES,
    VAMS_COLUMNS,
)

from tabs.add_vams_data.excel_writer.formatting import apply_table_formatting

from utils.file_handler import list_excel_sheets, validate_file
from utils.memory import memory_session, release_large_objects
from gui.qt_dialogs import DialogService


class AddVamsDataTab(ctk.CTkFrame):

    def __init__(
        self,
        master,
        app_state,
        logger,
    ):

        super().__init__(master)

        self.state = app_state

        self.logger = logger

        self.output_file = ctk.StringVar()

        self.raw_file = ctk.StringVar()

        self.raw_sheet = ctk.StringVar()

        self.dialogs = DialogService()

        self._build()
        self._bind_validation()


    def _build(self):
        self.grid_columnconfigure(0, weight=1)
        backend_text = f"Dialog backend: {self.dialogs.backend_name} (PySide/PyQt when available)"
        ctk.CTkLabel(self, text=backend_text, text_color="#60a5fa").grid(row=0, column=0, sticky="w", padx=12, pady=(4, 0))
        self._file_row(1, "📄 Output file", self.output_file, self._browse_output)
        self._file_row(2, "📊 Raw VAMS file", self.raw_file, self._browse_raw)
        sheet_card = ctk.CTkFrame(self, corner_radius=12)
        sheet_card.grid(row=3, column=0, sticky="ew", padx=10, pady=8)
        ctk.CTkLabel(sheet_card, text="🧾 Raw VAMS sheet").grid(row=0, column=0, sticky="w", padx=12, pady=(8,4))
        self.raw_sheet_combo = ctk.CTkComboBox(sheet_card, values=[""], variable=self.raw_sheet)
        self.raw_sheet_combo.grid(row=1, column=0, sticky="w", padx=12, pady=(0,8))
        self.run_btn = ctk.CTkButton(self, text="▶ Run Assessment", command=self.run)
        self.run_btn.grid(row=4, column=0, sticky="e", padx=12, pady=10)
        self.run_btn.configure(state="disabled")

    def _file_row(self, row, label, var, browse):
        card = ctk.CTkFrame(self, corner_radius=12)
        card.grid(row=row, column=0, sticky="ew", padx=10, pady=8)
        card.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(card, text=label).grid(row=0, column=0, sticky="w", padx=12, pady=(8,4))
        ctk.CTkEntry(card, textvariable=var).grid(row=1, column=0, sticky="ew", padx=12, pady=(0,8))
        ctk.CTkButton(card, text="Browse", command=browse, width=110).grid(row=1, column=1, padx=12)

    def _bind_validation(self):
        for var in (self.output_file, self.raw_file, self.raw_sheet):
            var.trace_add("write", lambda *_: self._update_run_state())
        self._update_run_state()

    def _update_run_state(self):
        enabled = bool(self.output_file.get().strip() and self.raw_file.get().strip() and self.raw_sheet.get().strip())
        self.run_btn.configure(state="normal" if enabled else "disabled")

    def _browse_output(self):
        path = self.dialogs.open_excel_file(title="Select generated workbook", save_workbook_only=True)
        if not path:
            return
        self.output_file.set(path)    


    def _browse_raw(self):
        path = self.dialogs.open_excel_file(title="Select raw VAMS file")
        if not path:
            return
        self.raw_file.set(path)
        sheets = list_excel_sheets(path)
        self.raw_sheet_combo.configure(values=sheets or [""])
        self.raw_sheet.set(
            sheets[0]
        )    
    def _validate_inputs(self):
        validate_file(
            self.output_file.get()
        )
        validate_file(
            self.raw_file.get()
        )
        if not self.raw_sheet.get():
            raise ValueError(
                "Please select raw VAMS sheet"
            )     
        
    def _read_generated_sheet(
        self,
        path: str,
        *sheet_names: str,
    ):
        for sheet_name in sheet_names:
            try:
                return read_sheet_as_df(
                    path,
                    sheet_name,
                )
            except ValueError:
                continue
        return None
    
    def _write_vams_columns_only(
        self,
        path: str,
        sheet_name: str,
        enriched_df,
    ):
        self.logger.info(
            "Opening workbook for VAMS writeback: %s",
            sheet_name,
        )
        wb = load_workbook(path)
        ws = wb[sheet_name]
        header_row = 2
        data_start_row = 3
        # ---------------------------------------------------
        # READ HEADERS
        # ---------------------------------------------------
        headers = {}
        for col_idx in range(
            1,
            ws.max_column + 1,
        ):
            header_value = ws.cell(
                row=header_row,
                column=col_idx,
            ).value
            if header_value is None:
                continue
            headers[
                str(header_value).strip()
            ] = col_idx
        # ---------------------------------------------------
        # REQUIRED MATCHING COLUMNS
        # ---------------------------------------------------
        required_columns = [
            "Name",
        ]
        for col in required_columns:
            if col not in headers:
                self.logger.error(
                    "Missing required column in Excel sheet: %s",
                    col,
                )
                return
        # ---------------------------------------------------
        # BUILD EXCEL LOOKUP USING TIERED KEYS
        # ---------------------------------------------------
        excel_records = []
        row_numbers = []
        for excel_row in range(
            data_start_row,
            ws.max_row + 1,
        ):
            record = {}
            for col_name in [
                "Name",
                "Host / Image",
                "Host",
                "IP",
                "Port",
                "CVE",
                "Scanner ID",
                "Plugin ID",
                "QID",
            ]:
                col_idx = headers.get(col_name)
                record[col_name] = (
                    ws.cell(excel_row, col_idx).value
                    if col_idx is not None
                    else ""
                )
            excel_records.append(record)
            row_numbers.append(excel_row)

        excel_lookup = {}
        for idx, row_keys in enumerate(build_fast_keys(
            pd.DataFrame(excel_records)
        )):
            for meta in row_keys:
                excel_lookup.setdefault(
                    meta["key"],
                    row_numbers[idx],
                )
        # ---------------------------------------------------
        # NORMALIZE ENRICHED DF
        # ---------------------------------------------------
        df = (
            enriched_df
            .fillna("")
            .astype(str)
            .copy()
        )
        updated = 0
        # ---------------------------------------------------
        # WRITE VAMS VALUES
        # ---------------------------------------------------
        for idx, row in df.iterrows():
            try:
                row_keys = build_fast_keys(
                    pd.DataFrame([row.to_dict()])
                )[0]
                excel_row = None
                for meta in row_keys:
                    excel_row = excel_lookup.get(
                        meta["key"]
                    )
                    if excel_row:
                        break
                if not excel_row:
                    continue
                row_updated = False
                # ---------------------------------------------
                # WRITE ONLY VAMS COLUMNS
                # ---------------------------------------------
                for col_name in (
                    VAMS_COLUMNS
                ):
                    col_idx = headers.get(
                        col_name
                    )
                    if col_idx is None:
                        continue
                    value = str(
                        row.get(
                            col_name,
                            "",
                        )
                    ).strip()
                    if not value:
                        continue
                    existing_cell = str(
                        ws.cell(
                            row=excel_row,
                            column=col_idx,
                        ).value or ""
                    ).strip().lower()
                    if existing_cell in [
                        "",
                        "nan",
                    ]:
                        ws.cell(
                            row=excel_row,
                            column=col_idx,
                            value=value,
                        )
                        row_updated = True
                if row_updated:
                    updated += 1
            except Exception:
                continue
        # ---------------------------------------------------
        # FORMAT SHEET
        # ---------------------------------------------------
        apply_table_formatting(ws)
        # ---------------------------------------------------
        # SAVE WORKBOOK
        # ---------------------------------------------------
        self.logger.info(
            "Saving workbook..."
        )
        wb.save(path)
        wb.close()
        self.logger.info(
            "Updated %s rows in %s",
            updated,
            sheet_name,
        )
    def _refresh_dashboard_charts(
        self,
        path: str,
        vams_chart_df,
    ):
        # -------------------------------------------------
        # READ SHEETS
        # -------------------------------------------------
        total_df = self._read_generated_sheet(
            path,
            *TOTAL_SHEET_CANDIDATES,
        )
        unique_df = self._read_generated_sheet(
            path,
            *UNIQUE_SHEET_CANDIDATES,
        )
        if (
            total_df is None
            or unique_df is None
        ):
            self.logger.warning(
                "Dashboard refresh skipped because required sheets were not found"
            )
            return
        # -------------------------------------------------
        # LOAD WORKBOOK
        # -------------------------------------------------
        wb = load_workbook(path)
        # -------------------------------------------------
        # GET DASHBOARD SHEET
        # -------------------------------------------------
        if "Dashboard" in wb.sheetnames:
            ws = wb["Dashboard"]
        elif "Summary Data" in wb.sheetnames:
            ws = wb["Summary Data"]
            ws.title = "Dashboard"
        else:
            ws = wb.create_sheet(
                "Dashboard",
                0,
            )
        # -------------------------------------------------
        # CLEAR EXISTING DASHBOARD
        # -------------------------------------------------
        ws.delete_rows(
            1,
            ws.max_row,
        )
        # -------------------------------------------------
        # REBUILD DASHBOARD
        # -------------------------------------------------
        write_summary_sheet(
            ws,
            total_df,
            total_df.iloc[0:0],
            unique_df,
            self.state["selected_project"],
            self.state["selected_scanner"],
            include_old_summary=False,
            include_vams_charts=True,
            vams_chart_df=vams_chart_df,
        )
        # -------------------------------------------------
        # SAVE
        # -------------------------------------------------
        wb.save(path)
        wb.close()
        self.logger.info(
            "Dashboard refreshed successfully"
        )

    def run(self):

        try:
            hooks = self.state.get("ui_hooks", {})
            hooks.get("set_run_state", lambda *_: None)("Running")
            hooks.get("set_stage", lambda *_: None)("Validate Inputs", 1)

            self.run_btn.configure(
                state="disabled"
            )
            mem_ctx = memory_session(self.logger, "TAB3 Add VAMS Data")
            mem_ctx.__enter__()

            self._validate_inputs()

            self.logger.info(
                "Loading workbook sheets..."
            )

            total_df = (
                self._read_generated_sheet(
                    self.output_file.get(),
                    *TOTAL_SHEET_CANDIDATES,
                )
            )

            unique_df = (
                self._read_generated_sheet(
                    self.output_file.get(),
                    *UNIQUE_SHEET_CANDIDATES,
                )
            )

            if total_df is None:

                raise ValueError(
                    "Could not find Total Vulnerabilities sheet"
                )

            if unique_df is None:

                raise ValueError(
                    "Could not find Unique Vulnerabilities sheet"
                )

            # -------------------------------------------------
            # PARSE INCOMING VAMS FILE
            # -------------------------------------------------

            self.logger.info(
                "Parsing incoming VAMS file..."
            )
            hooks.get("set_stage", lambda *_: None)("Parse", 2)

            incoming_vams = parse_scan_file(

                self.raw_file.get(),

                self.raw_sheet.get(),

                self.state["selected_scanner"],

                self.state["selected_project"],

                require_rows_after_filter=False,

                apply_severity_filter=False,

            ).df

            if incoming_vams.empty:

                self.dialogs.show_warning(
                    "Warning",
                    "No rows found in VAMS file",
                )

                return

            # -------------------------------------------------
            # INITIALIZE FAST ENGINE
            # -------------------------------------------------

            self.logger.info(
                "Initializing enrichment engine..."
            )
            hooks.get("set_stage", lambda *_: None)("Enrich", 4)

            engine = (
                FastVamsEnrichmentEngine()
            )

            engine.build_vams_lookup(
                incoming_vams
            )

            # -------------------------------------------------
            # SPECIAL FLOW:
            # 3UK + QUALYS
            # -------------------------------------------------

            is_3uk_qualys = (

                self.state["selected_project"]
                .strip()
                .casefold()
                == "3uk"

                and

                self.state["selected_scanner"]
                .strip()
                .casefold()
                == "qualys"
            )

            # -------------------------------------------------
            # ENRICH UNIQUE DF
            # -------------------------------------------------

            self.logger.info(
                "Enriching Unique Vulnerabilities..."
            )

            enriched_unique_df = (
                engine.enrich(
                    unique_df
                )
            )
            hooks.get("update_metrics", lambda **_: None)(
                unique_vulns=len(enriched_unique_df),
            )

            # -------------------------------------------------
            # AVAILABLE SHEETS
            # -------------------------------------------------

            available_sheets = (
                list_excel_sheets(
                    self.output_file.get()
                )
            )

            target_unique_sheet = next(

                (
                    sheet_name

                    for sheet_name
                    in UNIQUE_SHEET_CANDIDATES

                    if sheet_name
                    in available_sheets
                ),

                "Unique Vulnerabilities",
            )

            target_total_sheet = next(

                (
                    sheet_name

                    for sheet_name
                    in TOTAL_SHEET_CANDIDATES

                    if sheet_name
                    in available_sheets
                ),

                "Total Vulnerabilities",
            )

            # -------------------------------------------------
            # GENERIC FLOW
            # ALL SCANNERS EXCEPT:
            # 3UK + QUALYS
            # -------------------------------------------------

            if not is_3uk_qualys:

                self.logger.info(
                    "Generic flow detected."
                )

                # ---------------------------------------------
                # ENRICH TOTAL
                # ---------------------------------------------

                self.logger.info(
                    "Enriching Total Vulnerabilities..."
                )

                enriched_total_df = (
                    engine.enrich(
                        total_df
                    )
                )

                # ---------------------------------------------
                # WRITE TOTAL
                # ---------------------------------------------

                self.logger.info(
                    "Writing VAMS data into Total Vulnerabilities..."
                )

                self._write_vams_columns_only(

                    self.output_file.get(),

                    target_total_sheet,

                    enriched_total_df,
                )

                # ---------------------------------------------
                # WRITE UNIQUE
                # ---------------------------------------------

                self.logger.info(
                    "Writing VAMS data into Unique Vulnerabilities..."
                )

                self._write_vams_columns_only(

                    self.output_file.get(),

                    target_unique_sheet,

                    enriched_unique_df,
                )

            # -------------------------------------------------
            # SPECIAL FLOW:
            # 3UK + QUALYS
            # -------------------------------------------------

            else:

                self.logger.info(
                    "3UK + Qualys detected."
                )

                self.logger.info(
                    "Skipping Total Vulnerabilities write."
                )

                self.logger.info(
                    "Writing ONLY Unique Vulnerabilities..."
                )

                # ---------------------------------------------
                # ONLY UNIQUE
                # ---------------------------------------------

                self._write_vams_columns_only(

                    self.output_file.get(),

                    target_unique_sheet,

                    enriched_unique_df,
                )
            

            # -------------------------------------------------
            # REFRESH DASHBOARD
            # -------------------------------------------------

            self.logger.info(
                "Refreshing dashboard..."
            )
            hooks.get("set_stage", lambda *_: None)("Write", 5)

            self._refresh_dashboard_charts(

                self.output_file.get(),

                enriched_unique_df,
            )

            # -------------------------------------------------
            # COMPLETE
            # -------------------------------------------------

            output = self.output_file.get()

            self.logger.info(
                "VAMS merge completed: %s",
                output,
            )

            self.state["last_output_file"] = output
            self.dialogs.show_info(
                "Success",
                "VAMS data merged successfully",
            )
            hooks.get("set_run_state", lambda *_: None)("Success")

        except Exception as exc:

            self.logger.exception(
                "Add VAMS Data failed"
            )

            self.dialogs.show_error(
                "Error",
                str(exc),
            )
            hooks.get("set_run_state", lambda *_: None)("Failed")

        finally:
            try:
                mem_ctx.__exit__(None, None, None)
            except Exception:
                pass
            release_large_objects(
                locals(),
                [
                    "total_df",
                    "unique_df",
                    "incoming_vams",
                    "enriched_unique_df",
                    "enriched_total_df",
                    "engine",
                    "output",
                ],
            )

            self._update_run_state()


    def reset(self):
        self.output_file.set("")
        self.raw_file.set("")
        self.raw_sheet.set("")
        if hasattr(self, "raw_sheet_combo"):
            self.raw_sheet_combo.configure(values=[""])
        if hasattr(self, "run_btn"):
            self.run_btn.configure(state="disabled")
