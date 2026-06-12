"""Workbook-level Excel output orchestration."""

from __future__ import annotations

from pathlib import Path

#from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import (
    get_column_letter,
)
#from openpyxl.utils.dataframe import dataframe_to_rows

from tabs.add_vams_data.excel_writer.sheets import (
    write_disposition_sheet,
    write_main_sheet,
    write_summary_sheet,
)

from tabs.add_vams_data.excel_writer.three_uk_qualys import (
    is_three_uk_qualys_project,
    #three_uk_qualys_total_view,
)

TOTAL_VULNERABILITIES_SHEET_NAME = "Total Vulnerabilities"

_TOTAL_VULNERABILITY_ALIASES = {
    "total vulnerabilities",
    "total vulnerability",
    "total vulnerability sheet",
    "total_vulnerability_sheet",
    "total vulenrability",
}


def _normalized_sheet_name(value: object) -> str:

    return " ".join(
        str(value)
        .strip()
        .casefold()
        .replace("_", " ")
        .split()
    )


def normalize_total_sheet_name(
    sheet_name: str,
) -> str:

    if (
        _normalized_sheet_name(sheet_name)
        in _TOTAL_VULNERABILITY_ALIASES
    ):

        return TOTAL_VULNERABILITIES_SHEET_NAME

    return sheet_name

def autofit_worksheet_columns(ws):

    for column_cells in ws.columns:

        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            try:

                value_length = len(
                    str(cell.value)
                )

                if value_length > max_length:

                    max_length = value_length

            except:
                pass

        adjusted_width = min(
            max(max_length + 3, 12),
            50,
        )

        ws.column_dimensions[
            column_letter
        ].width = adjusted_width


def write_output(
    path,
    new_df,
    old_df,
    unique_df,
    project,
    scanner,
    include_old_sheet: bool = True,
    new_sheet_name: str = "Total Vulnerabilities",
    include_dashboard_sheet: bool = True,
):
    """Create workbook with required sheet order and persist to disk."""

    Path(path).parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    new_sheet_name = normalize_total_sheet_name(
        new_sheet_name
    )

    wb = Workbook()

    # ---------------------------------------------------------
    # DASHBOARD
    # ---------------------------------------------------------

    if include_dashboard_sheet:

        ws_summary = wb.active

        ws_summary.title = "Dashboard"

        write_summary_sheet(
            ws_summary,
            new_df,
            old_df,
            unique_df,
            project,
            scanner,
            include_old_summary=include_old_sheet,
        )

    else:

        ws_first = wb.active

        ws_first.title = new_sheet_name

        write_main_sheet(
            ws_first,
            new_df,
            project,
            scanner=scanner,
        )

    # ---------------------------------------------------------
    # TOTAL VULNERABILITIES
    # ---------------------------------------------------------

    if include_dashboard_sheet:

        ws_new = wb.create_sheet(
            new_sheet_name
        )

        # -----------------------------------------------------
        # SPECIAL FLOW:
        # 3UK + QUALYS
        # -----------------------------------------------------
        
        if is_three_uk_qualys_project(
            project,
            scanner,
        ):
        
            # -------------------------------------------------
            # IMPORTANT:
            # new_df is ALREADY mapped
            # DO NOT REMAP AGAIN
            # -------------------------------------------------
        
            qualys_total_df = new_df.copy()
        
            # -------------------------------------------------
            # META ROW
            # -------------------------------------------------
        
            ws_new["A1"] = "Project Name:"
            ws_new["B1"] = project
        
            ws_new["D1"] = "Scanner:"
            ws_new["E1"] = scanner



            # -------------------------------------------------
            # ROW 1 FULL BLUE
            # -------------------------------------------------

            blue_fill = PatternFill(
                fill_type="solid",
                fgColor="1F497D",
            )

            white_font = Font(
                color="FFFFFF",
                bold=True,
                size=10,
                name="Calibri",
            )

            center_alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

            thin_border = Border(

                left=Side(
                    style="thin",
                    color="D9D9D9",
                ),

                right=Side(
                    style="thin",
                    color="D9D9D9",
                ),

                top=Side(
                    style="thin",
                    color="D9D9D9",
                ),

                bottom=Side(
                    style="thin",
                    color="D9D9D9",
                ),
            )

            total_columns = len(
                qualys_total_df.columns
            )

            for col_idx in range(
                1,
                total_columns + 1,
            ):

                cell = ws_new.cell(
                    row=1,
                    column=col_idx,
                )

                cell.fill = blue_fill

                cell.font = Font(
                    color="FFFFFF",
                    bold=True,
                    size=10,
                    name="Calibri",
                )

                cell.alignment = center_alignment

                cell.border = thin_border

            # -------------------------------------------------
            # HEADER ROW
            # -------------------------------------------------

            for col_idx, col_name in enumerate(
                qualys_total_df.columns,
                start=1,
            ):

                ws_new.cell(
                    row=2,
                    column=col_idx,
                    value=col_name,
                )

            # -------------------------------------------------
            # QUALYS HEADER COLORS
            # -------------------------------------------------

            green_fill = PatternFill(
                fill_type="solid",
                fgColor="92D050",
            )

            blue_fill = PatternFill(
                fill_type="solid",
                fgColor="1F497D",
            )

            white_font = Font(
                color="FFFFFF",
                bold=True,
            )

            center_alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

            blue_headers = {

                "Protocol",
                "FQDN",
                "SSL",
                "First Detected",
                "Last Detected",
                "Times Detected",
                "Date Last Fixed",
                "CVE ID",
                "Vendor Reference",
                "Bugtraq ID",
                "CVSS",
                "Criticality",
                "CVSS Base",
                "CVSS Temporal",
                "Product",
                "CVSS Environment",
                "CVSS3",
                "CVSS3 Base",
                "CVSS3 Temporal",
                "Threat",
                "Impact",
                "Solution",
                "Exploitability",
                "Associated Malware",
                "Results",
                "PCI Vuln",
                "Ticket State",
                "Instance",
                "Category",
            }

            for col_idx, col_name in enumerate(
                qualys_total_df.columns,
                start=1,
            ):

                cell = ws_new.cell(
                    row=2,
                    column=col_idx,
                )

                if col_name in blue_headers:

                    cell.fill = blue_fill

                else:

                    cell.fill = green_fill

                cell.font = white_font

                cell.alignment = center_alignment

            # -------------------------------------------------
            # DATA ROWS
            # -------------------------------------------------

            for row_idx, row in enumerate(
                qualys_total_df.itertuples(index=False),
                start=3,
            ):

                for col_idx, value in enumerate(
                    row,
                    start=1,
                ):

                    ws_new.cell(
                        row=row_idx,
                        column=col_idx,
                        value=value,
                    )

        else:

            write_main_sheet(
                ws_new,
                new_df,
                project,
                scanner=scanner,
            )

    # ---------------------------------------------------------
    # OLD DATA
    # ---------------------------------------------------------

    if include_old_sheet:

        ws_old = wb.create_sheet(
            "Old Data"
        )

        write_main_sheet(
            ws_old,
            old_df,
            project,
            scanner=scanner,
        )

    # ---------------------------------------------------------
    # UNIQUE VULNERABILITIES
    # ---------------------------------------------------------

    ws_unique = wb.create_sheet(
        "Unique Vulnerabilities"
    )

    # -----------------------------------------------------
    # SPECIAL FLOW:
    # 3UK + QUALYS
    # -----------------------------------------------------

    if is_three_uk_qualys_project(
        project,
        scanner,
    ):

        # -------------------------------------------------
        # META ROW
        # -------------------------------------------------

        ws_unique["A1"] = "Project Name:"
        ws_unique["B1"] = project

        ws_unique["D1"] = "Scanner:"
        ws_unique["E1"] = scanner


        # -------------------------------------------------
        # ROW 1 FULL BLUE
        # -------------------------------------------------

        blue_fill = PatternFill(
            fill_type="solid",
            fgColor="1F497D",
        )
        white_font = Font(
            color="FFFFFF",
            bold=True,
            size=10,
            name="Calibri",
        )
        center_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        thin_border = Border(
            left=Side(
                style="thin",
                color="D9D9D9",
            ),
            right=Side(
                style="thin",
                color="D9D9D9",
            ),
            top=Side(
                style="thin",
                color="D9D9D9",
            ),
            bottom=Side(
                style="thin",
                color="D9D9D9",
            ),
        )
        total_columns = len(
            unique_df.columns
        )
        for col_idx in range(
            1,
            total_columns + 1,
        ):
            cell = ws_unique.cell(
                row=1,
                column=col_idx,
            )
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # -------------------------------------------------
        # HEADER ROW
        # -------------------------------------------------

        for col_idx, col_name in enumerate(
            unique_df.columns,
            start=1,
        ):

            ws_unique.cell(
                row=2,
                column=col_idx,
                value=col_name,
            )

        # -------------------------------------------------
        # UNIQUE HEADER COLORS
        # -------------------------------------------------

        green_fill = PatternFill(
            fill_type="solid",
            fgColor="8CC63F",
        )

        blue_fill = PatternFill(
            fill_type="solid",
            fgColor="1F497D",
        )

        white_font = Font(
            color="FFFFFF",
            bold=True,
            size=10,
            name="Calibri",
        )

        center_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        for col_idx in range(
            1,
            len(unique_df.columns) + 1,
        ):

            cell = ws_unique.cell(
                row=2,
                column=col_idx,
            )

            if col_idx >= 14:

                cell.fill = blue_fill

            else:

                cell.fill = green_fill

            cell.font = white_font

            cell.alignment = center_alignment

            cell.border = thin_border

        # -------------------------------------------------
        # DATA ROWS
        # -------------------------------------------------

        for row_idx, row in enumerate(
            unique_df.itertuples(index=False),
            start=3,
        ):

            for col_idx, value in enumerate(
                row,
                start=1,
            ):

                ws_unique.cell(
                    row=row_idx,
                    column=col_idx,
                    value=value,
                )

    # -----------------------------------------------------
    # GENERIC FLOW
    # -----------------------------------------------------

    else:
        write_main_sheet(
            ws_unique,
            unique_df,
            project,
            scanner=scanner,
        )

    # ---------------------------------------------------------
    # DISPOSITION TEMPLATE
    # ---------------------------------------------------------

    ws_disposition = wb.create_sheet(
        "Disposition"
    )

    write_disposition_sheet(
        ws_disposition,
        project,
        scanner,
    )

    # ---------------------------------------------------------
    # SAVE
    # ---------------------------------------------------------

    autofit_worksheet_columns(
        ws_new
    )

    autofit_worksheet_columns(
        ws_unique
    )

    wb.save(path)

    return path