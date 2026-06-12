"""Workbook-level Excel output orchestration."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from tabs.generate_tracking.excel_writer.sheets import (
    write_disposition_sheet,
    write_main_sheet,
    write_summary_sheet,
)

TOTAL_VULNERABILITIES_SHEET_NAME = "Total Vulnerabilities"
NEW_VULNERABILITIES_SHEET_NAME = "New Vulnerabilities"
OLD_VULNERABILITIES_SHEET_NAME = "Old Vulnerabilities"
UNIQUE_VULNERABILITIES_SHEET_NAME = "Unique Vulnerabilities"
DASHBOARD_SHEET_NAME = "Dashboard"
DISPOSITION_SHEET_NAME = "Disposition"

_TOTAL_VULNERABILITY_ALIASES = {
    "total vulnerabilities",
    "total vulnerability",
    "total vulnerability sheet",
    "total_vulnerability_sheet",
    "total vulenrability",
}


BLUE_FILL = PatternFill(fill_type="solid", fgColor="1F497D")
GREEN_FILL = PatternFill(fill_type="solid", fgColor="92D050")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
BOLD_FONT = Font(bold=True, size=10, name="Calibri")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

QUALYS_BLUE_HEADERS = {
    "Protocol",
    "FQDN",
    "SSL",
    "First Detected",
    "Last Detected",
    "Times Detected",
    "Date Last Fixed",
    "CVE ID",
    "Vendor Reference",
    "Bugtraq ID",
    "CVSS",
    "Criticality",
    "CVSS Base",
    "CVSS Temporal",
    "Product",
    "CVSS Environment",
    "CVSS3",
    "CVSS3 Base",
    "CVSS3 Temporal",
    "Threat",
    "Impact",
    "Solution",
    "Exploitability",
    "Associated Malware",
    "Results",
    "PCI Vuln",
    "Ticket State",
    "Instance",
    "Category",
}


def _normalized_sheet_name(value: object) -> str:
    return " ".join(str(value).strip().casefold().replace("_", " ").split())


def normalize_total_sheet_name(sheet_name: str) -> str:
    if _normalized_sheet_name(sheet_name) in _TOTAL_VULNERABILITY_ALIASES:
        return TOTAL_VULNERABILITIES_SHEET_NAME
    return sheet_name


def autofit_worksheet_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max(max_length + 3, 12), 50)


def _coerce_dataframe(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    return df.fillna("")


def _dashboard_ready_df(df: pd.DataFrame) -> pd.DataFrame:
    """Return a dashboard-compatible DataFrame without changing sheet output."""
    dashboard_df = _coerce_dataframe(df).copy()
    if "Risk" not in dashboard_df.columns and "Criticality" in dashboard_df.columns:
        dashboard_df["Risk"] = dashboard_df["Criticality"].astype(str).str.title()
    return dashboard_df


def _style_row_one(ws, total_columns: int, project: str, scanner: str):
    ws["A1"] = "Project Name:"
    ws["B1"] = project
    ws["D1"] = "Scanner:"
    ws["E1"] = scanner

    for col_idx in range(1, total_columns + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = BLUE_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER


def _write_dataframe_sheet(
    ws,
    df: pd.DataFrame,
    project: str,
    scanner: str,
    *,
    blue_headers: set[str] | None = None,
    blue_from_column: int | None = None,
):
    """Write any DataFrame to a tracking workbook sheet using the app's two-row header."""
    df = _coerce_dataframe(df)
    columns = [str(column) for column in df.columns]
    total_columns = max(len(columns), 5)

    _style_row_one(ws, total_columns, project, scanner)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        use_blue = (blue_headers is not None and col_name in blue_headers) or (
            blue_from_column is not None and col_idx >= blue_from_column
        )
        cell.fill = BLUE_FILL if use_blue else GREEN_FILL
        cell.font = WHITE_FONT if use_blue else BOLD_FONT
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_idx, row in enumerate(df.itertuples(index=False), start=3):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER

    ws.freeze_panes = "A3"
    autofit_worksheet_columns(ws)


def _write_tracking_sheet(ws, df: pd.DataFrame, project: str, scanner: str, *, sheet_kind: str):
    """Write Total/New/Old/Unique Generate Tracking sheets in template layout."""
    write_main_sheet(ws, _coerce_dataframe(df), project, scanner=scanner)


def write_output(
    path,
    new_df,
    old_df,
    unique_df,
    project,
    scanner,
    include_old_sheet: bool = True,
    new_sheet_name: str = TOTAL_VULNERABILITIES_SHEET_NAME,
    include_dashboard_sheet: bool = True,
    total_df: pd.DataFrame | None = None,
):
    """Create the Generate Tracking workbook with dashboard, data, and reference sheets."""
    Path(path).parent.mkdir(parents=True, exist_ok=True)

    total_sheet_name = normalize_total_sheet_name(new_sheet_name)
    total_df = _coerce_dataframe(total_df if total_df is not None else new_df)
    new_df = _coerce_dataframe(new_df)
    old_df = _coerce_dataframe(old_df)
    unique_df = _coerce_dataframe(unique_df)

    wb = Workbook()

    if include_dashboard_sheet:
        ws_dashboard = wb.active
        ws_dashboard.title = DASHBOARD_SHEET_NAME
        write_summary_sheet(
            ws_dashboard,
            _dashboard_ready_df(total_df),
            _dashboard_ready_df(old_df),
            _dashboard_ready_df(unique_df),
            project,
            scanner,
            include_old_summary=include_old_sheet,
        )
    else:
        ws_total = wb.active
        ws_total.title = total_sheet_name
        _write_tracking_sheet(ws_total, total_df, project, scanner, sheet_kind="total")

    if include_dashboard_sheet:
        ws_total = wb.create_sheet(total_sheet_name)
        _write_tracking_sheet(ws_total, total_df, project, scanner, sheet_kind="total")

    ws_new = wb.create_sheet(NEW_VULNERABILITIES_SHEET_NAME)
    _write_tracking_sheet(ws_new, new_df, project, scanner, sheet_kind="new")

    if include_old_sheet:
        ws_old = wb.create_sheet(OLD_VULNERABILITIES_SHEET_NAME)
        _write_tracking_sheet(ws_old, old_df, project, scanner, sheet_kind="old")

    ws_unique = wb.create_sheet(UNIQUE_VULNERABILITIES_SHEET_NAME)
    _write_tracking_sheet(ws_unique, unique_df, project, scanner, sheet_kind="unique")

    ws_disposition = wb.create_sheet(DISPOSITION_SHEET_NAME)
    write_disposition_sheet(ws_disposition, project, scanner)

    wb.save(path)
    return path
