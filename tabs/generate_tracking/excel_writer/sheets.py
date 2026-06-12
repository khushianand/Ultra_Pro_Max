"""Excel sheet-level writers for data, dashboard, and disposition sheets."""

from __future__ import annotations
import pandas as pd
from openpyxl.chart import BarChart3D, PieChart3D, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

from tabs.generate_tracking.excel_writer.formatting import (
    BLACK_THIN_BORDER,
    BOLD,
    BLUE,
    CENTER,
    DATA_ALIGNMENT,
    WHITE,
    auto_width,
    write_headers,
)
from visuals.highlight_logic import severity_fill
from tabs.generate_tracking.parser import TEMPLATE_COLUMNS
from visuals.summary.summary_generator import disposition_summary, expert_severity_summary, severity_chart_summary

DASHBOARD_SOURCE_START_COL = 27  # AA: off-screen chart source data, keeping visible dashboard chart-only.
DASHBOARD_VISIBLE_MAX_COL = 21  # A:U visible dashboard canvas.
DASHBOARD_VISIBLE_MAX_ROW = 39
DASHBOARD_BG = "EAF2FF"
DASHBOARD_HEADER = "1D4ED8"
DASHBOARD_PANEL = DASHBOARD_BG
SEVERITY_CHART_COLORS = ["9B0F06", "D53E0F", "F77F00", "FCBF49"]
BAR_CHART_COLORS = ["4472C4", "C00000", "F4B183", "70AD47", "7030A0", "8064A2", "92D050", "00B0F0"]
WRAP_CENTER = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
)


DISPOSITION_ORDER = [
    "Transferred",
    "Mitigate in Future Release",
    "Mitigate",
    "In Analysis",
    "False Positive/Not Relevant",
    "Accept/No Solution Expected",
    "Accept",
    "False Positive",
]



def _write_data_rows(ws, df: pd.DataFrame):
    for row_idx, row in enumerate(df[TEMPLATE_COLUMNS].itertuples(index=False), start=3):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        risk_cell = ws.cell(row=row_idx, column=4)
        risk_cell.fill = severity_fill(risk_cell.value)
        risk_cell.font = BOLD


def write_main_sheet(
    ws,
    df: pd.DataFrame,
    project: str,
    scanner: str,
):

    write_headers(ws)

    ws["B1"] = project
    ws["E1"] = scanner

    _write_data_rows(ws, df)

    ws.freeze_panes = "A3"

    auto_width(ws)#def write_main_sheet(ws, df: pd.DataFrame, project: str, scanner: str):


    # Use the Qualys-style Total Data layout only for the exact requested
    # combination: project 3UK + scanner Qualys + Total Data sheet. Every other
    # project/scanner/sheet combination falls through to the universal template below.

    

def _clear_dashboard(ws):
    """Clear dashboard contents defensively before rebuilding charts.

    Workbooks that have been edited/opened by Excel can contain stale merged-cell
    metadata.  In that state, openpyxl may raise ``KeyError`` while unmerging a
    range because one of the tracked merged cells is already missing from
    ``ws._cells``.  The dashboard is about to be fully redrawn, so stale merge
    metadata can be discarded safely instead of failing the whole VAMS workflow.
    """
    ws._charts = []
    for merged_range in list(ws.merged_cells.ranges):
        try:
            ws.unmerge_cells(str(merged_range))
        except KeyError:
            try:
                ws.merged_cells.ranges.remove(merged_range)
            except (KeyError, ValueError):
                pass
        except ValueError:
            # The range may already have been removed by openpyxl; continue so
            # the dashboard can still be rebuilt from a clean visible canvas.
            pass
    for row in ws.iter_rows():
        for cell in row:
            try:
                cell.value = None
            except AttributeError:
                # A stale MergedCell can be read-only if its merge metadata was
                # inconsistent. It is harmless because the sheet is redrawn below.
                pass

def _hide_chart_source_columns(ws, start_col: int, width: int):
    """Hide only the off-screen reference-content table used by charts.

    The dashboard charts must remain visible while their supporting source data
    stays out of sight.  Columns AA onward are dedicated to those references, so
    only those columns are hidden and the visible chart canvas remains untouched.
    """
    for col_idx in range(start_col, start_col + width):
        col = ws.column_dimensions[get_column_letter(col_idx)]
        col.hidden = True
        col.width = 0


def _prepare_chart_canvas(ws, project: str = "", scanner: str = ""):
    """Paint a polished visible Dashboard canvas behind the charts.

    Edit this function when changing Dashboard colors/layout. Chart data still
    lives in hidden AA onward reference columns; only the visible canvas is
    styled here.
    """
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None

    for col_idx in range(1, DASHBOARD_SOURCE_START_COL):
        col = ws.column_dimensions[get_column_letter(col_idx)]
        col.hidden = False
        col.width = 11

    for row_idx in range(1, DASHBOARD_VISIBLE_MAX_ROW + 1):
        ws.row_dimensions[row_idx].hidden = False
        ws.row_dimensions[row_idx].height = 21

    bg_fill = PatternFill("solid", fgColor=DASHBOARD_BG)
    for row in ws.iter_rows(
        min_row=1,
        max_row=DASHBOARD_VISIBLE_MAX_ROW,
        min_col=1,
        max_col=DASHBOARD_VISIBLE_MAX_COL,
    ):
        for cell in row:
            cell.fill = bg_fill

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=DASHBOARD_VISIBLE_MAX_COL)
    title = ws.cell(row=1, column=1)
    title.value = f"Vulnerability Summary Dashboard  |  Project: {project or '-'}  |  Scanner: {scanner or '-'}"
    title.fill = PatternFill("solid", fgColor=DASHBOARD_HEADER)
    title.font = Font(color="FFFFFF", bold=True, size=16)
    title.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx in (1, 2):
        ws.row_dimensions[row_idx].height = 26

    _paint_dashboard_panel(ws, 4, 1, 16, 7)
    _paint_dashboard_panel(ws, 4, 9, 16, 15)
    _paint_dashboard_panel(ws, 19, 1, 34, 7)
    _paint_dashboard_panel(ws, 19, 9, 37, 20)


def _paint_dashboard_panel(ws, start_row: int, start_col: int, end_row: int, end_col: int):
    """Paint a no-border sky-blue chart area behind Dashboard charts."""
    fill = PatternFill("solid", fgColor=DASHBOARD_PANEL)
    no_border = Border()
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.fill = fill
            cell.border = no_border

def _write_chart_source(ws, df: pd.DataFrame, start_row: int, start_col: int) -> int:
    for col_offset, col_name in enumerate(df.columns):
        ws.cell(start_row, start_col + col_offset, col_name)
    for row_offset, row in enumerate(df.itertuples(index=False), start=1):
        for col_offset, value in enumerate(row):
            ws.cell(start_row + row_offset, start_col + col_offset, value)
    return len(df)



def _color_series_points(series, colors):
    series.data_points = []
    for idx, color in enumerate(colors):
        point = DataPoint(idx=idx)
        point.graphicalProperties.solidFill = color
        series.data_points.append(point)


def _style_chart(chart, *, height: float, width: float):
    """Apply consistent dashboard chart sizing and rendering defaults."""
    chart.height = height
    chart.width = width
    chart.style = 26
    chart.plotVisOnly = False
    chart.visible_cells_only = False
    chart.display_blanks = "zero"
    chart.graphical_properties = GraphicalProperties(solidFill=DASHBOARD_PANEL, noFill=False)


def _add_pie(ws, title: str, source_row: int, source_col: int, size: int, anchor: str):
    chart = PieChart3D()
    chart.title = title
    chart.firstSliceAng = 270
    chart.legend.position = "r"
    data = Reference(ws, min_col=source_col + 1, min_row=source_row, max_row=source_row + size)
    cats = Reference(ws, min_col=source_col, min_row=source_row + 1, max_row=source_row + size)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    if chart.series:
        _color_series_points(chart.series[0], SEVERITY_CHART_COLORS[:size])
        chart.series[0].dLbls = DataLabelList()
        chart.series[0].dLbls.showVal = True
        chart.series[0].dLbls.showPercent = True
        chart.series[0].dLbls.showCatName = True
        chart.series[0].dLbls.separator = "\n"
    _style_chart(chart, height=8.9, width=15.2)
    ws.add_chart(chart, anchor)


def _add_bar(ws, title: str, source_row: int, source_col: int, rows: int, cols: int, anchor: str, *, height: float = 8.9, width: float = 15.2):
    chart = BarChart3D()
    chart.type = "col"
    chart.grouping = "standard"
    chart.shape = "box"
    chart.title = title
    chart.y_axis.title = None
    chart.x_axis.title = "Category"
    chart.legend.position = "b"
    data = Reference(ws, min_col=source_col + 1, max_col=source_col + cols - 1, min_row=source_row, max_row=source_row + rows)
    cats = Reference(ws, min_col=source_col, min_row=source_row + 1, max_row=source_row + rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    for idx, series in enumerate(chart.series):
        color = BAR_CHART_COLORS[idx % len(BAR_CHART_COLORS)]
        series.graphicalProperties.solidFill = color
        series.graphicalProperties.line.solidFill = "FFFFFF"
        series.dLbls = DataLabelList()
        series.dLbls.showVal = True
    if len(chart.series) == 1 and rows > 1:
        _color_series_points(chart.series[0], BAR_CHART_COLORS[:rows])
    _style_chart(chart, height=height, width=width)
    ws.add_chart(chart, anchor)

def _append_pie_charts(ws, total_df: pd.DataFrame, unique_df: pd.DataFrame, source_col: int = DASHBOARD_SOURCE_START_COL) -> int:
    """Place the two severity pies in the left-hand 2x2 dashboard grid."""
    unique_table = severity_chart_summary(unique_df, include_total=False)
    total_table = severity_chart_summary(total_df, include_total=False)

    unique_row = 1
    total_row = unique_row + len(unique_table) + 2
    unique_size = _write_chart_source(ws, unique_table, unique_row, source_col)
    total_size = _write_chart_source(ws, total_table, total_row, source_col)

    _add_pie(ws, "Unique Vulnerabilities per severity", unique_row, source_col, unique_size, "A4")
    _add_pie(ws, "Total Vulnerabilities per severity", total_row, source_col, total_size, "A19")
    return total_row + total_size + 2


def _append_vams_bar_charts(ws, vams_df: pd.DataFrame, source_row: int, source_col: int = DASHBOARD_SOURCE_START_COL):
    reported_expert = expert_severity_summary(vams_df)
    disposition = disposition_summary(vams_df, DISPOSITION_ORDER)

    reported_row = source_row
    disposition_row = reported_row + len(reported_expert) + 3
    reported_rows = _write_chart_source(ws, reported_expert, reported_row, source_col)
    disposition_rows = _write_chart_source(ws, disposition, disposition_row, source_col)

    _add_bar(
        ws,
        "Reported Severity Vs Expert Severity",
        reported_row,
        source_col,
        reported_rows,
        len(reported_expert.columns),
        "I4",
    )
    _add_bar(
        ws,
        "Disposition",
        disposition_row,
        source_col,
        disposition_rows,
        len(disposition.columns),
        "I19",
        height=10,
        width=20,
    )


def write_summary_sheet(
    ws,
    new_df,
    old_df,
    unique_df,
    project,
    scanner,
    include_old_summary: bool = True,
    include_vams_charts: bool = False,
    vams_chart_df: pd.DataFrame | None = None,
):
    """Write the chart-only Dashboard sheet.

    Render exactly four charts in a clean 2x2 layout. Supporting chart-source
    tables are written off-screen and hidden so the visible Dashboard stays clean.
    """
    _clear_dashboard(ws)
    _prepare_chart_canvas(ws, project, scanner)
    next_source_row = _append_pie_charts(ws, new_df, unique_df)
    _append_vams_bar_charts(
        ws,
        vams_chart_df if vams_chart_df is not None else new_df,
        next_source_row,
    )
    _hide_chart_source_columns(ws, DASHBOARD_SOURCE_START_COL, 3)


def write_disposition_sheet(ws, project: str = "", scanner: str = "", *_, **__):
    """Create static Disposition sheet template included in every output workbook."""

    # Row 1 project/scanner values
    ws["A1"] = "Project Name:"
    ws["D1"] = "Scanner:"
    ws["B1"] = project
    ws["E1"] = scanner

    # Apply row-1 color coding/styling across the entire first row.
    for col_idx in range(1, len(TEMPLATE_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = BLUE
        cell.font = WHITE
        cell.alignment = CENTER
        cell.border = BLACK_THIN_BORDER

    headers = ["Disposition Value", "Description"]
    for col_idx, title in enumerate(headers, start=1):
        # Table starts at row 3 (row 2 intentionally left blank)
        cell = ws.cell(row=3, column=col_idx, value=title)
        cell.fill = BLUE
        cell.font = WHITE
        cell.alignment = CENTER
        cell.border = BLACK_THIN_BORDER

    # 👉 EDIT HERE (COLUMN A VALUES):
    # Each tuple is: ("Disposition Value", "Description")
    # - First value goes to column A (Disposition Value).
    # - Second value goes to column B (Description).
    # Example:
    # ("Mitigate", "Fix planned in next quarterly release.")
    disposition_rows = [
        (
            "Transferred",
            "Responsibility for resolution is assumed by Customer (mainly used for scans). This is used for configuration or certificate related issues which typically need to be addressed by the customer.",
        ),
        (
            "Mitigate in Future Release",
            "Vulnerability will be mitigated but not in current product/release where vulnerability was reported (i.e. if not mitigated in this product/release, it will be fixed in a later release).",
        ),
        ("Mitigate", "Vulnerability will be fixed in current product/release configured in VAMS."),
        ("In Analysis", "Assessment in progress - intermediate state before final disposition value."),
        (
            "False Positive/Not Relevant",
            "False Positive could represent a known scanner issue or Not Relevant notification against 3rd party component used in product. False Positive should be used only when there is zero risk that the vulnerability exists.",
        ),
        (
            "Accept/No Solution Expected",
            "Risk is accepted. Used for situations where vulnerability risk is low and impact to release may be high, or when vendor states they will not provide a patch/fix.",
        ),
    ]

    for row_idx, (value, description) in enumerate(disposition_rows, start=4):
        ws.cell(row=row_idx, column=1, value=value)
        ws.cell(row=row_idx, column=2, value=description)

    # 👉 OPTIONAL: adjust row heights if you want larger cells for long descriptions.
    # Example: ws.row_dimensions[3].height = 70
    #ws.row_dimensions[4].height = 60
    #ws.row_dimensions[6].height = 45
    table_start_row = 3
    table_end_row = 3 + len(disposition_rows)
    for row in ws.iter_rows():
        for cell in row:
            in_disposition_table = (
                table_start_row <= cell.row <= table_end_row
                and 1 <= cell.column <= 2
            )
            cell.border = BLACK_THIN_BORDER if in_disposition_table else None
            if cell.row in [1, 3]:
                cell.alignment = WRAP_CENTER
            else:
                cell.alignment = DATA_ALIGNMENT

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 15

    auto_width(ws)
