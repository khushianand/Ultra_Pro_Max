"""Severity-to-color mapping used for Excel cell highlighting."""

from openpyxl.styles import PatternFill

SEVERITY_COLORS = {
    "Critical": "009B0F06",
    "High": "00D53E0F",
    "Medium": "00F77F00",
    "Low": "00FCBF49",
}


def severity_fill(severity: str) -> PatternFill:
    color = SEVERITY_COLORS.get(str(severity).strip().title())
    if not color:
        return PatternFill(fill_type=None)
    return PatternFill(start_color=color, end_color=color, fill_type="solid")
