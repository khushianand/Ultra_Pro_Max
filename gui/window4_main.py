"""Modern cybersecurity dashboard layout for workflow tabs + logs."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
import os
import subprocess
import sys
import tkinter as tk

import customtkinter as ctk
from openpyxl import load_workbook

from tabs.add_vams_data.add_vams_data import AddVamsDataTab
from tabs.generate_tracking.generate_tracking import GenerateTrackingTab
from tabs.make_new_report.make_new_report import MakeNewReportTab
from gui.ui.cards import SummaryCards
from gui.ui.header import HeaderPanel
from gui.ui.logs_panel import LogsPanel
from gui.ui.sidebar import Sidebar
from gui.ui.themes import palette
from gui.ui.workflow_stepper import WorkflowStepper
from gui.qt_dialogs import DialogService


class Window4Main(ctk.CTkFrame):
    """Enterprise-style dashboard shell hosting existing processing tabs."""

    TAB_NAMES = ["📄 Make Report", "📊 Generate Tracking", "🛡 Add VAMS Data", "📈 Show Summary"]
    STAGE_TO_STEPPER = {
        "Validate Inputs": "Inputs",
        "Parse": "Parse",
        "Compare": "Compare",
        "Compare/Aggregate": "Compare",
        "Enrich": "Enrich",
        "Write": "Write",
        "Write/Format": "Write",
    }

    def __init__(self, master, state, logger, on_start_again=None):
        super().__init__(master)
        self.state = state
        self.logger = logger
        self.on_start_again = on_start_again
        self.theme_name = self.state.get("theme_name", "Dark")
        self._stage_progress = 0
        self._timer_job = None
        self.dialogs = DialogService()
        ctk.set_appearance_mode("light" if self.theme_name == "Light" else "dark")
        self.colors = palette(self.theme_name)
        self._apply_window_background()
        self.configure(fg_color=self.colors["bg"])
        self._build()

    def _build(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self.header = HeaderPanel(self, self.colors, lambda: self.theme_name)
        self.header.grid(row=0, column=0, sticky="ew", padx=12, pady=(10, 8))

        self.paned = tk.PanedWindow(
            self,
            orient=tk.VERTICAL,
            sashwidth=8,
            sashrelief="raised",
            bd=0,
            bg=self.colors["bg"],
            showhandle=True,
        )
        self.paned.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 4))

        body = ctk.CTkFrame(self.paned, fg_color="transparent")
        body.grid_columnconfigure(0, weight=0)
        body.grid_columnconfigure(1, weight=1)
        body.grid_rowconfigure(0, weight=1)

        self.sidebar = Sidebar(body, self.colors, self._on_nav)
        self.sidebar.grid(row=0, column=0, sticky="ns", padx=(0, 8), pady=(0, 8))

        center = ctk.CTkFrame(
            body,
            corner_radius=12,
            fg_color=self.colors.get("center", self.colors["panel"]),
            border_width=1,
            border_color=self.colors["border"],
        )
        center.grid(row=0, column=1, sticky="nsew", pady=(0, 8))
        center.grid_rowconfigure(3, weight=1)
        center.grid_columnconfigure(0, weight=1)

        self.cards = SummaryCards(center, self.colors)
        self.cards.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 8))

        self.stepper = WorkflowStepper(center, self.colors)
        self.stepper.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 8))

        controls = ctk.CTkFrame(center, fg_color="transparent")
        controls.grid(row=2, column=0, sticky="ew", padx=10, pady=(0, 6))
        ctk.CTkButton(controls, text="⚙ Theme", fg_color=self.colors.get("button_blue", self.colors["secondary"]), command=self._toggle_theme).pack(side="left", padx=(0, 8))
        ctk.CTkButton(controls, text="▣ Export Logs", fg_color=self.colors.get("button_blue", self.colors["secondary"]), command=self._export_logs).pack(side="left", padx=(0, 8))
        ctk.CTkButton(controls, text="⟳ Reset", fg_color=self.colors.get("danger", self.colors.get("red", "#EF4444")), command=self._reset_current_tab).pack(side="left")

        tabs_holder = ctk.CTkFrame(center, fg_color="transparent")
        tabs_holder.grid(row=3, column=0, sticky="nsew", padx=10, pady=(0, 10))
        tabs_holder.grid_rowconfigure(0, weight=1)
        tabs_holder.grid_columnconfigure(0, weight=1)

        self.tabview = ctk.CTkTabview(tabs_holder, command=self._on_tab_change)
        self.tabview.grid(row=0, column=0, sticky="nsew")
        for name in self.TAB_NAMES:
            self.tabview.add(name)
        self.tabs = [
            MakeNewReportTab(self.tabview.tab("📄 Make Report"), self.state, self.logger),
            GenerateTrackingTab(self.tabview.tab("📊 Generate Tracking"), self.state, self.logger),
            AddVamsDataTab(self.tabview.tab("🛡 Add VAMS Data"), self.state, self.logger),
        ]
        for tab in self.tabs:
            tab.pack(fill="both", expand=True, padx=6, pady=6)
        self._build_summary_tab(self.tabview.tab("📈 Show Summary"))

        self.logs = LogsPanel(self.paned, self.colors, self._open_output_file)
        self.paned.add(body, minsize=360)
        self.paned.add(self.logs, minsize=130)

        self.status_bar = ctk.CTkLabel(
            self,
            text=f"● Ready | Project: {self.state.get('selected_project','-')} | Scanner: {self.state.get('selected_scanner','-')} | Mode: {self.state.get('entry_mode','-')} | v1.0.0",
            fg_color=self.colors.get("version_bg", self.colors["panel"]),
            text_color=self.colors.get("primary", self.colors["text"]),
            corner_radius=14,
            pady=6,
        )
        self.status_bar.grid(row=2, column=0, sticky="ew", padx=12, pady=(0, 8))

        self.state["ui_hooks"] = {
            "set_stage": self.set_stage,
            "set_run_state": self.set_run_state,
            "update_metrics": self.update_metrics,
        }
        metrics = self.state.get("live_metrics")
        if metrics:
            metrics.subscribe(lambda m: self.cards.update_metrics(
                **{
                    "Vulnerabilities Processed": m.total_vulns,
                    "Unique Vulnerabilities": m.unique_vulns,
                    "Processing Time": f"{m.processing_time}s",
                    "Success Rate": f"{m.success_rate}%",
                }
            ))
            metrics.notify()

    def _build_summary_tab(self, tab):
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(0, weight=1)
        panel = ctk.CTkFrame(tab, corner_radius=12, fg_color=self.colors["panel"], border_width=1, border_color=self.colors["border"])
        panel.grid(row=0, column=0, sticky="nsew", padx=12, pady=12)
        ctk.CTkLabel(
            panel,
            text="📈 Show Summary Dashboard",
            font=ctk.CTkFont(size=22, weight="bold"),
            text_color=self.colors["text"],
        ).pack(anchor="w", padx=18, pady=(18, 8))
        ctk.CTkLabel(
            panel,
            text="Open the latest output workbook directly on its Dashboard / Summary sheet.",
            text_color=self.colors.get("muted", self.colors["text"]),
        ).pack(anchor="w", padx=18, pady=(0, 16))
        ctk.CTkButton(panel, text="Open Summary Dashboard", command=self._open_summary_dashboard).pack(anchor="w", padx=18, pady=(0, 18))

    def _on_tab_change(self):
        if self.tabview.get() == "📈 Show Summary":
            self.after(100, self._open_summary_dashboard)

    def _on_nav(self, name: str):
        mapping = {
            "Start Again": -3,
            "Make Report": 0,
            "Generate Tracking": 1,
            "Add VAMS Data": 2,
            "Show Summary": 3,
            "Logs": -1,
            "Settings": -2,
        }
        idx = mapping.get(name)
        if idx is None:
            return
        if idx >= 0:
            self.tabview.set(self.TAB_NAMES[idx])
            if idx == 3:
                self.after(100, self._open_summary_dashboard)
            return
        if idx == -1:
            self.logs.text.focus_set()
            self.logger.info("Logs panel focused from sidebar")
            return
        if idx == -3:
            self.logger.info("Start Again selected from sidebar")
            if self.on_start_again is not None:
                self.after_idle(self.on_start_again)
            return
        self._open_settings_modal()

    def _apply_window_background(self):
        bg = self.colors["bg"]
        for widget in (self, self.master, self.winfo_toplevel()):
            try:
                widget.configure(bg=bg)
            except Exception:
                try:
                    widget.configure(fg_color=bg)
                except Exception:
                    pass

    def _toggle_theme(self):
        order = ["Dark", "Light", "Cybersecurity Neon"]
        self.theme_name = order[(order.index(self.theme_name) + 1) % len(order)]
        self.state["theme_name"] = self.theme_name
        ctk.set_appearance_mode("light" if self.theme_name == "Light" else "dark")
        self.colors = palette(self.theme_name)
        self._apply_window_background()
        self.configure(fg_color=self.colors["bg"])
        if self._timer_job:
            self.after_cancel(self._timer_job)
            self._timer_job = None
        self.dialogs = DialogService()
        for child in self.winfo_children():
            child.destroy()
        self._build()
        self.logger.info("Theme switched to %s", self.theme_name)

    def _current_output_path(self) -> str:
        current_name = self.tabview.get() if hasattr(self, "tabview") else ""
        index = {"📄 Make Report": 0, "📊 Generate Tracking": 1, "🛡 Add VAMS Data": 2}.get(current_name)
        if index is not None and index < len(self.tabs):
            output_var = getattr(self.tabs[index], "output_file", None)
            if output_var is not None and output_var.get().strip():
                return output_var.get().strip()
        return str(self.state.get("last_output_file", "")).strip()

    def _open_output_file(self):
        output_path = self._current_output_path()
        if not output_path:
            self.dialogs.show_warning("Open Output File", "No output file is selected or generated yet.")
            return

        path = Path(output_path)
        if not path.exists():
            self.dialogs.show_warning("Open Output File", f"Output file was not found:\n{path}")
            return

        if sys.platform.startswith("win"):
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(path)])
        else:
            subprocess.Popen(["xdg-open", str(path)])
        self.logger.info("Opened output file: %s", path)

    def _open_summary_dashboard(self):
        output_path = self._current_output_path()
        if not output_path:
            self.dialogs.show_warning("Show Summary", "No output file is selected or generated yet.")
            return
        path = Path(output_path)
        if not path.exists():
            self.dialogs.show_warning("Show Summary", f"Output file was not found:\n{path}")
            return
        try:
            wb = load_workbook(path)
            for sheet_name in ("Dashboard", "Summary Data"):
                if sheet_name in wb.sheetnames:
                    wb.active = wb.sheetnames.index(sheet_name)
                    break
            wb.save(path)
            wb.close()
        except Exception as exc:
            self.logger.warning("Could not activate dashboard sheet before opening output: %s", exc)
        self._open_output_file()

    def _open_settings_modal(self):
        modal = ctk.CTkToplevel(self)
        modal.title("Settings")
        modal.geometry("420x220")
        modal.grab_set()
        ctk.CTkLabel(modal, text="Settings", font=ctk.CTkFont(size=20, weight="bold")).pack(anchor="w", padx=16, pady=(16, 8))
        ctk.CTkLabel(modal, text=f"Current Theme: {self.theme_name}").pack(anchor="w", padx=16, pady=4)
        ctk.CTkButton(modal, text="Switch Theme", command=self._toggle_theme).pack(anchor="w", padx=16, pady=8)
        ctk.CTkLabel(modal, text="Use sidebar Logs to jump to the live console panel.").pack(anchor="w", padx=16, pady=8)

    def _export_logs(self):
        path = self.dialogs.save_file(
            title="Export logs",
            default_extension=".log",
            filetypes=[("Log Files", "*.log"), ("Text", "*.txt")],
            qt_filter="Log Files (*.log);;Text Files (*.txt)",
            initialfile=f"run-log-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.log",
        )
        if not path:
            return
        with open(path, "w", encoding="utf-8") as f:
            f.write(self.logs.text.get("1.0", "end"))
        self.logger.info("Logs exported to %s", path)

    def _reset_current_tab(self):
        current_name = self.tabview.get()
        index = {"📄 Make Report": 0, "📊 Generate Tracking": 1, "🛡 Add VAMS Data": 2}.get(current_name)
        if index is not None:
            current = self.tabs[index]
            if hasattr(current, "reset"):
                current.reset()
        metrics = self.state.get("live_metrics")
        if metrics and hasattr(metrics, "reset"):
            metrics.reset()
        self._stage_progress = 0
        self.stepper.set_active("")
        self.header.status_var.set("Ready")
        self.logs.text.delete("1.0", "end")
        self.logger.info("Reset executed")

    def _schedule_timer_tick(self):
        if self._timer_job:
            self.after_cancel(self._timer_job)
        self._timer_job = self.after(1000, self._timer_tick)

    def _timer_tick(self):
        self._timer_job = None
        metrics = self.state.get("live_metrics")
        if metrics and getattr(metrics, "_started_at", None) is not None:
            metrics.tick()
            self._schedule_timer_tick()

    @staticmethod
    def _normalized_excel_name(value: object) -> str:
        return " ".join(str(value or "").strip().casefold().replace("_", " ").split())

    def _find_workbook_sheet(self, wb, sheet_names: tuple[str, ...]) -> str | None:
        requested = {self._normalized_excel_name(name) for name in sheet_names}
        for actual_name in wb.sheetnames:
            if self._normalized_excel_name(actual_name) in requested:
                return actual_name
        return None

    def _count_sheet_name_entries(self, path: str, sheet_names: tuple[str, ...]) -> int:
        """Count non-empty Name-column entries in one of the requested output sheets."""
        workbook_path = Path(path)
        if not workbook_path.exists():
            return 0

        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            sheet_name = self._find_workbook_sheet(wb, sheet_names)
            if not sheet_name:
                return 0

            ws = wb[sheet_name]
            header_row = 2
            name_aliases = {"name", "title", "vulnerability", "plugin name"}
            name_col_idx = None
            for cell in ws[header_row]:
                if self._normalized_excel_name(cell.value) in name_aliases:
                    name_col_idx = cell.column
                    break

            if name_col_idx is None:
                self.logger.warning("Could not find Name column in sheet: %s", sheet_name)
                return 0

            count = 0
            for row_idx in range(header_row + 1, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=name_col_idx).value
                if str(value or "").strip().casefold() not in {"", "nan", "none"}:
                    count += 1
            return count
        finally:
            wb.close()

    def _refresh_output_metrics(self):
        output_path = self._current_output_path()
        if not output_path:
            return
        total_count = self._count_sheet_name_entries(
            output_path,
            ("Total Vulnerabilities", "Total_Vulnerabilities", "Total Vulnerability", "Total Data", "New Data"),
        )
        unique_count = self._count_sheet_name_entries(
            output_path,
            ("Unique Vulnerabilities", "Unique_Vulnerabilities", "Unique Data"),
        )
        self.update_metrics(total_vulns=total_count, unique_vulns=unique_count)

    def set_run_state(self, value: str):
        metrics = self.state.get("live_metrics")
        if metrics:
            if value == "Running":
                metrics.start()
                self._stage_progress = 0
                self._schedule_timer_tick()
            else:
                if self._timer_job:
                    self.after_cancel(self._timer_job)
                    self._timer_job = None
                if value == "Success":
                    self._refresh_output_metrics()
                metrics.stop(success=value == "Success")
        self.header.status_var.set(value)

    def set_stage(self, stage_name: str, progress_value: int):
        self.stepper.set_active(self.STAGE_TO_STEPPER.get(stage_name, "Inputs"))
        metrics = self.state.get("live_metrics")
        if metrics:
            self._stage_progress = max(0, min(int(progress_value), 5))
            metrics.success_rate = min(99, int((self._stage_progress / 5) * 100))
            metrics.notify()
        self.logger.info("Stage: %s", stage_name)

    def update_metrics(self, **kwargs):
        metrics = self.state.get("live_metrics")
        if not metrics:
            return
        aliases = {"duplicates_removed": "unique_vulns", "enriched_records": "unique_vulns"}
        for k, v in kwargs.items():
            attr = aliases.get(k, k)
            if hasattr(metrics, attr):
                setattr(metrics, attr, v)
        metrics.notify()

    def attach_log_handler(self):
        from utils.logger import UILogHandler
        import logging

        handler = UILogHandler(self.append_log)
        handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s", "%H:%M:%S"))
        self.logger.addHandler(handler)

    def append_log(self, msg: str):
        self.logs.append(msg)
